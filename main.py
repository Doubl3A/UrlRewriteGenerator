from urllib.parse import urlparse
from xml.dom import minidom
from openpyxl.reader.excel import load_workbook
from tldextract import tldextract
from helpers.ColorHelper import theme_and_tint_to_rgb
import config

# Initial workbook setup
wb = load_workbook(config.file, data_only=True)
sheet = wb.active
max_column: int = sheet.max_row

# Get cell indexes. Filter on matching cell color if enabled
cell_indexes = []
if config.enable_cell_color_filter:
    for row in range(1, max_column):
        # Get cell color
        loc = f'A{row}'
        cell = sheet[loc]
        theme = cell.fill.start_color.theme
        tint = cell.fill.start_color.tint
        color = theme_and_tint_to_rgb(wb, theme, tint)

        # Check for desired bg color
        if color in config.colors:
            cell_indexes.append(row)
else:
    cell_indexes = range(1, max_column)

print(f'A total of {cell_indexes.__len__()} matching cells were found')

# Xml file setup
root = minidom.Document()
xml = root.createElement('rewrite')
rules = root.createElement('rules')

# Error rows
error_rows = []

# Add UrlRewrite
for row in cell_indexes:

    # If initialization of a var fails, the cell value probably isn't a valid URL
    try:
        # Config
        old_url: str = sheet[f'A{row}'].value.replace("www.", "")
        parsed_old_url = urlparse(old_url)
        new_url: str = sheet[f'B{row}'].value.replace("www.", "")
        parsed_new_url = urlparse(new_url)

        name = f"{parsed_new_url.hostname.replace('.', '-')} {parsed_old_url.path}"

        match_url = parsed_old_url.path.split("/", 1)[1]
        add_pattern = tldextract.extract(old_url).registered_domain
        action_url = new_url.replace("www.", "", 1)
    except:
        error_rows.append(row)
        continue

    # Create a new rule
    rule = root.createElement('rule')
    rule.setAttribute('name', name)
    rule.setAttribute('stopProcessing', 'true')
    rule.setAttribute('patternSyntax', 'ECMAScript')

    # Create rule match
    match = root.createElement('match')
    match.setAttribute('url', match_url)

    rule.appendChild(match)

    # Create condition
    condition = root.createElement('conditions')
    # Condition add
    add = root.createElement('add')
    add.setAttribute('input', '{HTTP_HOST}')
    add.setAttribute('pattern', add_pattern)
    add.setAttribute('ignoreCase', 'true')

    condition.appendChild(add)
    rule.appendChild(condition)

    # Create action
    action = root.createElement('action')
    action.setAttribute('type', 'Redirect')
    action.setAttribute('url', action_url)
    action.setAttribute('redirectType', 'Permanent')

    rule.appendChild(action)

    # Add new rule to rewrite node
    rules.appendChild(rule)

# Append all rules to root
xml.appendChild(rules)
root.appendChild(xml)

# Xml string
xml_str = root.toprettyxml(indent='\t')

# Write to file
with open(f'{config.rewrite_file_name}.xml', 'w') as f:
    f.write(xml_str)
print(f'Created a rewrite file with name: {config.rewrite_file_name}')

if error_rows.__len__() != 0:
    print(f'Failed to initiate config vars for {error_rows.__len__()} rows. Rows that failed to initiate: {error_rows}')
